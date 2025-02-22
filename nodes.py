import openpyxl
import os

class ExcelPicker:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": os.path.join("TESTLIST.xlsx")}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_number": ("INT", {"default": 1, "min": 1, "max": 10000}),
                "prefix": ("STRING", {"default": ""}),
            },
        }

    # Only three outputs: Column A, Column B, and Column C
    RETURN_TYPES = ("STRING", "STRING", "STRING")
    RETURN_NAMES = ("Column_A", "Column_B", "Column_C")
    FUNCTION = "pick_prompt"
    CATEGORY = "ExcelPicker"

    def pick_prompt(self, excel_path, sheet_name, row_number, prefix):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_excel_path = os.path.join(current_dir, excel_path)
        
        if not os.path.exists(full_excel_path):
            raise FileNotFoundError(f"Excel file not found: {full_excel_path}")
        
        # Load workbook in default mode (not read_only) so that max_row is computed correctly
        workbook = openpyxl.load_workbook(full_excel_path)
        
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")
        
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        if max_row is None or row_number > max_row:
            workbook.close()
            raise ValueError(f"Row {row_number} exceeds the number of rows ({max_row}) in the sheet or the sheet is empty.")
        
        # Retrieve the specified row as a list of cells
        row_cells = list(sheet[row_number])
        workbook.close()
        
        # Retrieve values from the first three cells (Columns A, B, and C)
        values = []
        for i in range(3):
            if i < len(row_cells):
                cell_value = row_cells[i].value
                if cell_value is None:
                    cell_value = ""
            else:
                cell_value = ""
            values.append(str(cell_value))
        
        # Prepend prefix if provided
        if prefix.strip():
            values = [prefix.strip() + " " + v for v in values]
        
        return tuple(values)
